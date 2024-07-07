import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .models import Budget, BudgetItem
from projects.models import Contractor, Client, Project
from decimal import Decimal
from django.conf import settings

def copy_images(src_sheet, dest_sheet):
    """Copiar imágenes de una hoja de trabajo a otra."""
    for image in src_sheet._images:
        dest_sheet.add_image(image, image.anchor)

def export_budget_report_to_excel(request, pk):
    budget = get_object_or_404(Budget, pk=pk)
    project = budget.project
    client = project.client
    contractor = project.contractor
    
    # Cargar la plantilla de Excel
    file_path = f'{settings.BASE_DIR}/static/plantilla.xlsx'
    plantilla = openpyxl.load_workbook(file_path)
    hoja_cotizacion = plantilla['COTIZACION']

    # Llenar los datos del contratista
    hoja_cotizacion['C13'] = contractor.contractor_name
    hoja_cotizacion['C14'] = contractor.contractor_ruc
    hoja_cotizacion['C15'] = contractor.address

    # Llenar los datos del cliente
    hoja_cotizacion['C18'] = client.client_name
    hoja_cotizacion['C20'] = client.email
    hoja_cotizacion['C21'] = client.phone

    # Llenar los datos del proyecto
    hoja_cotizacion['C27'] = project.project_name
    hoja_cotizacion['D35'] = project.start_date
    hoja_cotizacion['D36'] = project.end_date
    hoja_cotizacion['G30'] = project.total_budget

    # Crear la hoja de presupuesto
    ws_budget = plantilla.create_sheet(title="Presupuesto")
    
    # Copiar las imágenes de la hoja de cotización a la nueva hoja de presupuesto
    copy_images(hoja_cotizacion, ws_budget)
    
    # Título del presupuesto
    ws_budget['A1'] = f'PRESUPUESTO: {budget.project.project_name}'
    ws_budget['A1'].font = Font(bold=True, size=14)

    # Encabezados de la tabla
    headers = ['Código de Ítem', 'Descripción de Ítem', 'Unidad', 'Cantidad', 'Precio Diario', 'Precio Total de Días']
    ws_budget.append(headers)
    for cell in ws_budget[2]:  # Segunda fila
        cell.font = Font(bold=True)
    
    # Datos de los ítems
    total_parcial = Decimal('0')
    for item in budget.items.all():
        precio_total_dias = item.precio_total_diario * budget.dias
        total_parcial += precio_total_dias
        ws_budget.append([
            item.inventario.num_articulo,
            item.inventario.descripcion,
            item.inventario.unidad_medida,
            item.cantidad,
            item.inventario.precio_unitario_diario,
            precio_total_dias
        ])

    # Ajustar ancho de las columnas
    for column in ws_budget.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 2
        ws_budget.column_dimensions[column[0].column_letter].width = adjusted_width

    # Cálculos adicionales
    gastos_administrativos = total_parcial * Decimal('0.1')
    utilidad = total_parcial * Decimal('0.1')
    total = total_parcial + gastos_administrativos + utilidad
    total_en_dolares = total / Decimal('3.75')  # Suponiendo un tipo de cambio fijo

    # Añadir el resumen financiero
    ws_budget.append([''])
    ws_budget.append(['TOTAL PARCIAL', '', '', '', '', f'S/ {total_parcial:.2f}'])
    ws_budget.append(['GASTOS ADMINISTRATIVOS 10%', '', '', '', f'0.1', f'S/ {gastos_administrativos:.2f}'])
    ws_budget.append(['UTILIDAD 10%', '', '', '', f'0.1', f'S/ {utilidad:.2f}'])
    ws_budget.append(['TOTAL', '', '', '', '', f'S/ {total:.2f}'])
    ws_budget.append(['TOTAL EN DÓLARES', '', '', '', '', f'$ {total_en_dolares:.2f}'])
    
    # Resaltar la fila de TOTAL EN DÓLARES
    last_row = ws_budget.max_row
    for cell in ws_budget[last_row]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Mover la hoja de presupuesto a la segunda posición
    # Crear la respuesta HTTP con el nombre del archivo según el nombre del proyecto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cotizacion_{project.project_name}.xlsx"'
    plantilla.save(response)
    return response
