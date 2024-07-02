from django.db.models import Sum
from datetime import datetime
from decimal import Decimal, InvalidOperation
import logging
import re
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views import View
from django.db import transaction
from django.template.loader import render_to_string
from django.utils.dateparse import parse_date
from django.utils import timezone
from openpyxl import load_workbook
import pdfplumber
from rest_framework import generics
from reportes.models import Proveedor
from .models import (FacturaElectronica, OrdenDeCompra, OrdenVenta, 
                     ItemOrdenVenta, CobrosOrdenVenta)
from .serializers import (OrdenVentaSerializer, ItemOrdenVentaSerializer)
from .forms import (CantidadPedidaForm, FacturaElectronicaForm, ItemOrdenVentaForm, 
                    OrdenVentaForm, OrdenVentaUploadForm, CobrosOrdenVentaForm)
from django.views.generic.edit import CreateView
from django.views.generic import ListView
from django.contrib import messages


# TODO: Función revisada
class OrdenVentaCRUDView(View):
    template_name = "ordenventa_crud.html"

    def get(self, request, *args, **kwargs):
        # Obtén la acción de la URL (solo delete ya que edit se eliminó)
        action = request.GET.get("action", None)

        # Lógica para eliminar
        if action == "delete":
            orden_venta_id = request.GET.get("delete", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                orden_venta.delete()
                return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))

        # Lógica por defecto para mostrar la lista y el formulario de creación
        form = OrdenVentaForm()
        ordenes_venta = OrdenVenta.objects.all().order_by("-id")
        return render(
            request, self.template_name, {"ordenes_venta": ordenes_venta, "form": form}
        )

    def post(self, request, *args, **kwargs):
        form = OrdenVentaForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))
        else:
            ordenes_venta = OrdenVenta.objects.all().order_by("-id")
            return render(
                request,
                self.template_name,
                {"ordenes_venta": ordenes_venta, "form": form},
            )

# TODO: Funcion revisada
def procesar_orden_venta_excel(request, ordenventa_id):
    if request.method == "POST":
        if "submit-excel" in request.POST:  # Identificar si se envió el formulario de Excel
            form_excel = OrdenVentaUploadForm(request.POST, request.FILES)
            if form_excel.is_valid():
                archivo_excel = request.FILES["archivo_excel"]
                workbook = load_workbook(archivo_excel)
                sheet = workbook.active

                # Verificar que la orden de venta existe
                orden_venta = get_object_or_404(OrdenVenta, pk=ordenventa_id)

                # Nombres de las columnas
                col_nro_articulo = 'Número de artículo'
                col_desc_articulo = 'Descripción del artículo'
                col_cantidad = 'Cantidad'
                col_precio_bruto = 'Precio bruto'
                col_total_bruto = 'Total bruto (ML)'  # o 'Total (ML)', ajustar según sea necesario

                # Encuentra los índices de las columnas
                header = [cell.value for cell in sheet[1]]
                idx_nro_articulo = header.index(col_nro_articulo)
                idx_desc_articulo = header.index(col_desc_articulo)
                idx_cantidad = header.index(col_cantidad)
                idx_precio_bruto = header.index(col_precio_bruto)
                idx_total_bruto = header.index(col_total_bruto)

                def limpiar_valor(valor):
                    if isinstance(valor, str):
                        # Eliminar el símbolo de moneda y convertir a float
                        valor = re.sub(r'[^\d.,]', '', valor)
                        try:
                            return Decimal(valor.replace(',', ''))
                        except ValueError:
                            return None
                    return valor

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[idx_nro_articulo] is None or row[idx_desc_articulo] is None or row[idx_cantidad] is None:
                        continue

                    cantidad = limpiar_valor(row[idx_cantidad])
                    precio_bruto = limpiar_valor(row[idx_precio_bruto])
                    total_bruto = limpiar_valor(row[idx_total_bruto])

                    if cantidad is None:
                        continue  # Si no se puede convertir la cantidad, saltar esta fila

                    ItemOrdenVenta.objects.create(
                        ordenventa=orden_venta,
                        nro_articulo=row[idx_nro_articulo],
                        desc_articulo=row[idx_desc_articulo],
                        cantidad=cantidad,
                        precio_bruto=precio_bruto,
                        total_bruto=total_bruto,
                    )
                return redirect("filtered_item_list", ordenventa_id=ordenventa_id)
            else:
                form_item = ItemOrdenVentaForm()  # Inicializar formulario de ítem vacío si no se procesa el de Excel
        elif "submit-item" in request.POST:  # Identificar si se envió el formulario de ítem individual
            form_item = ItemOrdenVentaForm(request.POST)
            form_excel = OrdenVentaUploadForm()  # Inicializar formulario de Excel vacío por defecto
            if form_item.is_valid():
                nuevo_item = form_item.save(commit=False)
                nuevo_item.ordenventa_id = ordenventa_id  # Asignar manualmente el ID de la orden de venta
                nuevo_item.save()
                return redirect("filtered_item_list", ordenventa_id=ordenventa_id)
    else:
        form_excel = OrdenVentaUploadForm()
        form_item = ItemOrdenVentaForm()

    return render(
        request,
        "formulario_orden_venta.html",
        {
            "form_excel": form_excel,
            "form_item": form_item,
            "ordenventa_id": ordenventa_id,
        },
    )


# TODO: Funcion revisada
class ItemListView(ListView):
    model = ItemOrdenVenta
    template_name = 'items/itemordenventa_list.html'
    context_object_name = 'items'

    def get_context_data(self, **kwargs):
        context = super(ItemListView, self).get_context_data(**kwargs)
        ordenventa_id = self.kwargs.get('ordenventa_id')
        if (ordenventa_id):
            ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
            context['ordenventa'] = ordenventa
            
            items = ItemOrdenVenta.objects.filter(ordenventa=ordenventa)
            for item in items:
                total_cantidad_ordenes_compra = item.ordenes_de_compra.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
                item.cantidad_restante = (item.cantidad or 0) - total_cantidad_ordenes_compra

            context['items'] = items
        return context

    def get_queryset(self):
        ordenventa_id = self.kwargs.get('ordenventa_id')
        if ordenventa_id:
            ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
            return ItemOrdenVenta.objects.filter(ordenventa=ordenventa)
        return super().get_queryset()

    def post(self, request, *args, **kwargs):
        ordenventa_id = self.kwargs.get('ordenventa_id')
        item_id = request.POST.get('item_id')
        item = ItemOrdenVenta.objects.get(pk=item_id)
        form = CantidadPedidaForm(request.POST, instance=item)

        if form.is_valid():
            form.instance.solicitar_nueva_orden = request.POST.get('solicitar_nueva_orden') == 'True'
            form.save()

            # Llamada a la actualización del bruto restante
            item.update_bruto_restante()

            if ordenventa_id:
                return redirect('filtered_item_list', ordenventa_id=ordenventa_id)
        return redirect('item_list')    


# TODO: Funcion revisada / puede ir a la app orden compra
# ORDEN AUTOMATICAS
def ver_ordenes_compra(request, ordenventa_id=None):
    if ordenventa_id:
        ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
        ordenes_compra = OrdenDeCompra.objects.filter(
            item_orden_venta__ordenventa=ordenventa,
        )
    else:
        ordenventa = None
        ordenes_compra = OrdenDeCompra.objects.all()  # Sin filtro, muestra todas las órdenes de compra

    proveedores = Proveedor.objects.all()

    return render(
        request,
        "ver_orden_compra.html",
        {
            "ordenes_compra": ordenes_compra,
            "ordenventa": ordenventa,
            "proveedores": proveedores,
        },
    )
# TODO: Funcion revisada / puede ir a la app orden compra
def actualizar_orden_de_compra(request, id):
    if request.method == "POST":
        orden_de_compra = get_object_or_404(OrdenDeCompra, pk=id)

        # Campos existentes
        cantidad = request.POST.get("cantidad", "")
        precio_actual = request.POST.get("precio_actual", "")
        igv = request.POST.get("igv", "")
        detraccion = request.POST.get("detraccion", "")

        # Nuevos campos
        clase = request.POST.get("clase", "")
        tipo_pago = request.POST.get("tipo_pago", "")
        proveedor_id = request.POST.get(
            "proveedor", ""
        )  # Cambiado a proveedor_id para claridad
        cuotas = request.POST.get("cuotas", "")
        fecha_pago = request.POST.get(
            "fecha_pago", ""
        )  # Nuevo campo para fecha de pago
        comprobante_pago = request.FILES.get(
            "comprobante_pago", None
        )  # Nuevo campo para el PDF

        # Asegurarse de convertir a tipos numéricos antes de realizar cálculos
        try:
            orden_de_compra.cantidad = (
                int(cantidad) if cantidad else orden_de_compra.cantidad
            )
            orden_de_compra.precio_actual = (
                float(precio_actual) if precio_actual else orden_de_compra.precio_actual
            )
            orden_de_compra.igv = float(igv) if igv else orden_de_compra.igv
            orden_de_compra.detraccion = (
                float(detraccion) if detraccion else orden_de_compra.detraccion
            )
            orden_de_compra.clase = clase if clase else orden_de_compra.clase
            orden_de_compra.cuotas = int(cuotas) if cuotas else orden_de_compra.cuotas

            # Actualización del proveedor
            if proveedor_id:
                proveedor = Proveedor.objects.filter(id=proveedor_id).first()
                orden_de_compra.proveedor = proveedor
            else:
                orden_de_compra.proveedor = None

            # Actualizar fecha_pago si se proporciona
            if fecha_pago:
                orden_de_compra.fecha_pago = parse_date(fecha_pago)

            # Actualizar comprobante_pago si se proporciona
            if comprobante_pago:
                orden_de_compra.comprobante_pago = comprobante_pago

            # Actuliazzar el tipo de pago
            if tipo_pago:  # Solo actualizar si se proporcionó un valor
                orden_de_compra.tipo_pago = tipo_pago



            orden_de_compra.save()
            # Aquí puedes decidir cómo responder después de actualizar la orden de compra.
            return HttpResponse(
                f"Orden de compra actualizada con éxito: {orden_de_compra.precio_total}"
            )

        except ValueError:
            # Manejo de error o devolver un mensaje al usuario
            return HttpResponse("Error en los datos proporcionados", status=400)

    else:
        # Método no permitido
        return HttpResponse("Método no permitido", status=405)

# TODO: Funcion revisada / puede ir a la app orden compra
def eliminar_orden_de_compra(request, id):
    if request.method == "POST":
        with transaction.atomic():
            orden_de_compra = get_object_or_404(OrdenDeCompra, pk=id)
            item_orden_venta = orden_de_compra.item_orden_venta
            ordenventa_id = item_orden_venta.ordenventa.id
            orden_de_compra.delete()

            # Actualizar el bruto restante después de eliminar la OrdenDeCompra
            item_orden_venta.update_bruto_restante()

            messages.success(request, "Orden de compra eliminada correctamente.")
            if ordenventa_id:
                return redirect('ver_ordenes_compra', ordenventa_id=ordenventa_id)
            else:
                return redirect('nombre_de_tu_url_para_listar_todas_ordenes')
    else:
        return HttpResponse("Método no permitido", status=405)


# TODO: Funcion revisada / puede ir a la app orden compra
# experimento :=
def ver_ordenes_pago(request, ordenventa_id=None):
    ordenes_de_pago = []
    ordenventa = None
    fecha = None

    if ordenventa_id:
        ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
        ordenes_de_pago = OrdenDeCompra.objects.filter(
            item_orden_venta__ordenventa=ordenventa,
            fecha_pago__isnull=False
        ).select_related("proveedor", "item_orden_venta__ordenventa")
    else:
        fecha_str = request.GET.get("fecha_pago")
        if fecha_str:
            fecha = parse_date(fecha_str)
            if fecha:
                ordenes_de_pago = OrdenDeCompra.objects.filter(
                    fecha_pago=fecha
                ).select_related("proveedor", "item_orden_venta__ordenventa")

    context = {
        "ordenes_de_pago": ordenes_de_pago,
        "ordenventa": ordenventa,
        "fecha": fecha,
    }
    return render(request, "ver_orden_pago.html", context)




def ver_cobros_orden_venta(request, ordenventa_id):
    orden_venta = get_object_or_404(OrdenVenta, id=ordenventa_id)
    cobros = CobrosOrdenVenta.objects.filter(orden_venta=orden_venta)

    return render(request, 'cobros/lista_cobros_orden_venta.html', {
        'orden_venta': orden_venta,
        'cobros': cobros
    })


## vista automatica de facturas okokokok
def ver_facturas_por_ordenventa(request, ordenventa_id):
    orden_venta = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    facturas = FacturaElectronica.objects.filter(orden_venta=orden_venta)
    context = {
        'orden_venta': orden_venta,
        'facturas': facturas,
    }
    return render(request, 'cobros/ver_facturas.html', context)

def eliminar_factura(request, factura_id):
    factura = get_object_or_404(FacturaElectronica, pk=factura_id)
    orden_venta_id = factura.orden_venta.id
    factura.delete()
    return redirect(reverse('ver_facturas_por_ordenventa', kwargs={'ordenventa_id': orden_venta_id}))

def editar_factura(request, factura_id):
    factura = get_object_or_404(FacturaElectronica, pk=factura_id)
    if request.method == 'POST':
        form = FacturaElectronicaForm(request.POST, instance=factura)
        if form.is_valid():
            form.save()
            # Si usas HTMX, puedes querer devolver un fragmento actualizado en lugar de hacer una redirección completa
            return HttpResponse(render_to_string('fragmentos/factura_actualizada.html', {'factura': factura}))
    else:
        form = FacturaElectronicaForm(instance=factura)

    return HttpResponse(render_to_string('fragmentos/formulario_editar_factura.html', {'form': form, 'factura_id': factura.id}))

def upload_pdf_cobro(request, ordenventa_id):
    orden_venta = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    if request.method == "POST":
        pdf_file = request.FILES.get('pdf_file')
        if pdf_file:
            texto_completo = ""
            with pdfplumber.open(pdf_file) as pdf:
                for pagina in pdf.pages:
                    texto_completo += pagina.extract_text() + "\n"

            # Aquí incluyes todo el código de extracción que has definido anteriormente
            # Asegúrate de ajustar el manejo de los datos según sea necesario
            # Por ejemplo, asegúrate de convertir las cadenas de fechas y decimales correctamente
            # Buscar la serie y el correlativo
            patron_serie_correlativo = r"E\d{3}-\d+"
            resultado_serie_correlativo = re.search(patron_serie_correlativo, texto_completo)
            serie_correlativo = resultado_serie_correlativo.group() if resultado_serie_correlativo else None
            logging.warning(texto_completo)

            # Buscar la fecha de emisión
            patron_fecha_emision = r"Fecha de Emisión\s*:\s*(\d{2}/\d{2}/\d{4})"
            resultado_fecha_emision = re.search(patron_fecha_emision, texto_completo)
            fecha_emision = resultado_fecha_emision.group(1) if resultado_fecha_emision else None

            # Buscar el cliente (Señor(es))
            patron_cliente = r"Señor\(es\) :\s*(.*)\n"
            resultado_cliente = re.search(patron_cliente, texto_completo)
            cliente = resultado_cliente.group(1) if resultado_cliente else None

            # Buscar el RUC del cliente
            patron_ruc_cliente = r"RUC :\s*(\d+)"
            resultado_ruc_cliente = re.search(patron_ruc_cliente, texto_completo)
            ruc_cliente = resultado_ruc_cliente.group(1) if resultado_ruc_cliente else None

            # Buscar el tipo de moneda
            patron_tipo_moneda = r"Tipo de Moneda\s*:\s*(.*)\n"
            resultado_tipo_moneda = re.search(patron_tipo_moneda, texto_completo)
            tipo_moneda = resultado_tipo_moneda.group(1) if resultado_tipo_moneda else None

            # Ajustar la búsqueda de la descripción para que termine antes de los detalles financieros
            # Suponiendo que "Segun OS" marca el final relevante de la descripción del servicio
            patron_descripcion = r"Descripción\s*(.*?)(?=Sub Total Ventas|$)"
            resultado_descripcion = re.search(patron_descripcion, texto_completo, re.DOTALL)
            descripcion = resultado_descripcion.group(1).strip() if resultado_descripcion else None

            # Buscar el importe total
            patron_importe_total = r"Importe Total\s*:\s*(\$|S/)(.*)\n"
            resultado_importe_total = re.search(patron_importe_total, texto_completo)
            importe_total = resultado_importe_total.group(2).replace(',', '') if resultado_importe_total else None
            logging.warning(f"Importe total (sin formato): {importe_total}")
            
            # Buscar el monto de la detracción
            patron_detraccion = r"Monto detracción:\s*(\$|S/)(.*)\n"
            resultado_detraccion = re.search(patron_detraccion, texto_completo)
            detraccion = resultado_detraccion.group(2).replace(',', '') if resultado_detraccion else None

            # Buscar el monto neto a cobrar
            patron_neto_cobrar = r"Monto neto pendiente de pago\s*:\s*(\$|S/)(.*)\n"
            resultado_neto_cobrar = re.search(patron_neto_cobrar, texto_completo)
            neto_cobrar = resultado_neto_cobrar.group(2).replace(',', '') if resultado_neto_cobrar else None

            # Buscar el total de cuotas
            patron_total_cuotas = r"Total de Cuotas\s*:\s*(\d+)"
            resultado_total_cuotas = re.search(patron_total_cuotas, texto_completo)
            total_cuotas = resultado_total_cuotas.group(1) if resultado_total_cuotas else None

            # Buscar la fecha de vencimiento de la cuota
            patron_fecha_vencimiento = r"Fec. Venc.\s*Monto\n\d+\s*(\d{2}/\d{2}/\d{4})"
            resultado_fecha_vencimiento = re.search(patron_fecha_vencimiento, texto_completo)
            fecha_vencimiento = resultado_fecha_vencimiento.group(1) if resultado_fecha_vencimiento else None

            # Solo intentar convertir la fecha si no es None
            if fecha_vencimiento:
                fecha_vencimiento_dt = datetime.strptime(fecha_vencimiento, "%d/%m/%Y")
                fecha_vencimiento_format = fecha_vencimiento_dt.strftime('%Y-%m-%d')
            else:
                fecha_vencimiento_dt = None
                fecha_vencimiento_format = None
                
            if fecha_emision:
                fecha_emision_format = datetime.strptime(fecha_emision, '%d/%m/%Y').strftime('%Y-%m-%d')
            else:
                fecha_emision_format = None

            logging.warning(f"neto a cobrar: {neto_cobrar}")
            logging.warning(f"Cuotas: {total_cuotas}")

            def convertir_decimal(cadena):
                try:
                    if cadena is None:
                        return None
                    cadena_limpia = cadena.strip()  # Eliminar espacios en blanco al inicio y al final
                    return Decimal(cadena_limpia)
                except InvalidOperation:
                    logging.error(f"Error al convertir {cadena} a Decimal")
                    return None

            # Convertir los valores a Decimal
            importe_total_decimal = convertir_decimal(importe_total)
            detraccion_decimal = convertir_decimal(detraccion)
            monto_neto_cobrar_decimal = convertir_decimal(neto_cobrar)
            monto_neto_cobrar_decimal = importe_total_decimal
            
            # Validar que los valores no sean None
            if importe_total_decimal is None:
                logging.error(f"Error al convertir importe_total: {importe_total}")
            if detraccion_decimal is None:
                logging.error(f"Error al convertir detraccion: {detraccion}")
            if monto_neto_cobrar_decimal is None:
                logging.error(f"Error al convertir neto_cobrar: {neto_cobrar}")

            # Crear instancia del modelo, incluyendo valores nulos donde corresponda
            factura = FacturaElectronica(
                orden_venta=orden_venta,
                serie_correlativo=serie_correlativo,
                fecha_emision=fecha_emision_format,
                cliente=cliente,
                ruc_cliente=ruc_cliente,
                tipo_moneda=tipo_moneda,
                descripcion=descripcion,
                importe_total=importe_total_decimal,
                detraccion=detraccion_decimal,
                monto_neto_cobrar=monto_neto_cobrar_decimal,
                total_cuotas=int(total_cuotas) if total_cuotas and total_cuotas.isdigit() else None,
                fecha_vencimiento=fecha_vencimiento_format,
            )
            factura.save()

            # Redirigir o mostrar un mensaje de éxito
            return redirect(reverse('ver_facturas_por_ordenventa', args=[ordenventa_id])) # Asegúrate de definir esta URL

    return render(request, 'cobros/upload_pdf_cobros.html')




















# desde aqui no hay nada interesante :P


# Vistas para OrdenVenta
class ListaOrdenesVenta(generics.ListCreateAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


class DetalleOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


# Vistas para ItemOrdenVenta
class ListaItemsOrdenVenta(generics.ListCreateAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer


class DetalleItemOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer
