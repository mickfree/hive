from datetime import datetime
from decimal import Decimal
import re
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views import View
from django.db import transaction
from django.template.loader import render_to_string
from django.utils.dateparse import parse_date
from openpyxl import load_workbook
import pdfplumber
from rest_framework import generics
from inventario.models import Inventario
from reportes.models import Proveedor
from .models import FacturaElectronica, OrdenDeCompra, OrdenVenta, ItemOrdenVenta, CobrosOrdenVenta
from .serializers import OrdenVentaSerializer, ItemOrdenVentaSerializer
from .forms import FacturaElectronicaForm, ItemOrdenVentaForm, OrdenVentaForm, OrdenVentaUploadForm, CobrosOrdenVentaForm
from django.views.generic.edit import CreateView
from django.views.generic import ListView
from django.shortcuts import render
from .models import OrdenDeCompra
from django.utils import timezone



# antiguos view debemos depurar despues
class OrdenVentaCRUDView(View):
    template_name = "ordenventa_crud.html"

    def get(self, request, *args, **kwargs):
        # Obtén la acción de la URL (edit o delete)
        action = request.GET.get("action", None)

        # Lógica para editar
        if action == "edit":
            orden_venta_id = request.GET.get("edit", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                form = OrdenVentaForm(instance=orden_venta)
            else:
                form = OrdenVentaForm()

        # Lógica para eliminar
        elif action == "delete":
            orden_venta_id = request.GET.get("delete", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                orden_venta.delete()
                return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))

        # Lógica por defecto para mostrar la lista y el formulario de creación
        else:
            form = OrdenVentaForm()

        ordenes_venta = OrdenVenta.objects.all().order_by("-id")
        return render(
            request, self.template_name, {"ordenes_venta": ordenes_venta, "form": form}
        )

    def post(self, request, *args, **kwargs):
        form = OrdenVentaForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))
        else:
            ordenes_venta = OrdenVenta.objects.all()
            return render(
                request,
                self.template_name,
                {"ordenes_venta": ordenes_venta, "form": form},
            )


# FUNCION PARA PROCESAR EL EXCEL Y ITEMS INDIVIDUALES
def procesar_orden_venta_excel(request, ordenventa_id):
    if request.method == "POST":
        if (
            "submit-excel" in request.POST
        ):  # Identificar si se envió el formulario de Excel
            form_excel = OrdenVentaUploadForm(request.POST, request.FILES)
            if form_excel.is_valid():
                archivo_excel = request.FILES["archivo_excel"]
                workbook = load_workbook(archivo_excel)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is None or row[1] is None or row[2] is None:
                        continue
                    ItemOrdenVenta.objects.create(
                        ordenventa_id=ordenventa_id,
                        nro_articulo=row[0],
                        desc_articulo=row[1],
                        cantidad=row[2],
                        precio_bruto=row[6],
                        total_bruto=row[12],
                    )
                return redirect("ver_items_orden_venta", ordenventa_id=ordenventa_id)
            else:
                form_item = (
                    ItemOrdenVentaForm()
                )  # Inicializar formulario de ítem vacío si no se procesa el de Excel
        elif (
            "submit-item" in request.POST
        ):  # Identificar si se envió el formulario de ítem individual
            form_item = ItemOrdenVentaForm(request.POST)
            form_excel = (
                OrdenVentaUploadForm()
            )  # Inicializar formulario de Excel vacío por defecto
            if form_item.is_valid():
                nuevo_item = form_item.save(commit=False)
                nuevo_item.ordenventa_id = (
                    ordenventa_id  # Asignar manualmente el ID de la orden de venta
                )
                nuevo_item.save()
                return redirect("ver_items_orden_venta", ordenventa_id=ordenventa_id)
    else:
        form_excel = OrdenVentaUploadForm()
        form_item = ItemOrdenVentaForm()

    return render(
        request,
        "formulario_orden_venta.html",
        {
            "form_excel": form_excel,
            "form_item": form_item,
            "ordenventa_id": ordenventa_id,
        },
    )


# VER LOS ITEMS SEGUN ID
def ver_items_orden_venta(request, ordenventa_id):
    ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    items = ItemOrdenVenta.objects.filter(ordenventa=ordenventa)

    # Verificar si la solicitud es una solicitud HTMX
    if request.headers.get("HX-Request", "false").lower() == "true":
        html = render_to_string("fragmentos/items_orden_venta.html", {"items": items})
        return JsonResponse({"html": html})

    return render(
        request,
        "ver_items_orden_venta.html",
        {"ordenventa": ordenventa, "items": items},
    )


# ORDEN AUTOMATICAS
def ver_ordenes_compra(request, ordenventa_id):
    ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    ordenes_compra = OrdenDeCompra.objects.filter(
        item_orden_venta__ordenventa=ordenventa,
        # Aquí puedes descomentar y agregar cualquier otro filtro que necesites.
        # item_orden_venta__enviado=True
    )
    # Aquí agregas la obtención de todos los proveedores disponibles.
    proveedores = Proveedor.objects.all()

    return render(
        request,
        "ver_orden_compra.html",
        {
            "ordenes_compra": ordenes_compra,
            "ordenventa": ordenventa,
            "proveedores": proveedores,  # Pasas la lista de proveedores a la plantilla.
        },
    )


# mas ordenes automaticas :P (y es la mas importante pelotudo no la friegues esta vez)
def actualizar_orden_de_compra(request, id):
    if request.method == "POST":
        orden_de_compra = get_object_or_404(OrdenDeCompra, pk=id)

        # Campos existentes
        cantidad = request.POST.get("cantidad", "")
        precio_actual = request.POST.get("precio_actual", "")
        igv = request.POST.get("igv", "")
        detraccion = request.POST.get("detraccion", "")

        # Nuevos campos
        clase = request.POST.get("clase", "")
        tipo_pago = request.POST.get("tipo_pago", "")
        proveedor_id = request.POST.get(
            "proveedor", ""
        )  # Cambiado a proveedor_id para claridad
        cuotas = request.POST.get("cuotas", "")
        fecha_pago = request.POST.get(
            "fecha_pago", ""
        )  # Nuevo campo para fecha de pago
        comprobante_pago = request.FILES.get(
            "comprobante_pago", None
        )  # Nuevo campo para el PDF

        # Asegurarse de convertir a tipos numéricos antes de realizar cálculos
        try:
            orden_de_compra.cantidad = (
                int(cantidad) if cantidad else orden_de_compra.cantidad
            )
            orden_de_compra.precio_actual = (
                float(precio_actual) if precio_actual else orden_de_compra.precio_actual
            )
            orden_de_compra.igv = float(igv) if igv else orden_de_compra.igv
            orden_de_compra.detraccion = (
                float(detraccion) if detraccion else orden_de_compra.detraccion
            )
            orden_de_compra.clase = clase if clase else orden_de_compra.clase
            orden_de_compra.cuotas = int(cuotas) if cuotas else orden_de_compra.cuotas

            # Actualización del proveedor
            if proveedor_id:
                proveedor = Proveedor.objects.filter(id=proveedor_id).first()
                orden_de_compra.proveedor = proveedor
            else:
                orden_de_compra.proveedor = None

            # Actualizar fecha_pago si se proporciona
            if fecha_pago:
                orden_de_compra.fecha_pago = parse_date(fecha_pago)

            # Actualizar comprobante_pago si se proporciona
            if comprobante_pago:
                orden_de_compra.comprobante_pago = comprobante_pago

            # Actuliazzar el tipo de pago
            if tipo_pago:  # Solo actualizar si se proporcionó un valor
                orden_de_compra.tipo_pago = tipo_pago

                

            orden_de_compra.save()
            # Aquí puedes decidir cómo responder después de actualizar la orden de compra.
            return HttpResponse(
                f"Orden de compra actualizada con éxito: {orden_de_compra.precio_total}"
            )

        except ValueError:
            # Manejo de error o devolver un mensaje al usuario
            return HttpResponse("Error en los datos proporcionados", status=400)

    else:
        # Método no permitido
        return HttpResponse("Método no permitido", status=405)


# ordenes d epago :) gracias totales


# experimento :=
def ver_ordenes_pago(request, ordenventa_id=None):
    ordenes_de_pago = []
    ordenventa = None
    fecha = None  # Inicializa fecha aquí

    if ordenventa_id:
        ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
        ordenes_de_pago = OrdenDeCompra.objects.filter(
            item_orden_venta__ordenventa=ordenventa, precio_total__gt=0.00
        ).select_related("proveedor")
    else:
        fecha_str = request.GET.get("fecha_pago")
        if fecha_str:
            fecha = parse_date(fecha_str)
            if fecha:
                ordenes_de_pago = OrdenDeCompra.objects.filter(
                    fecha_pago=fecha, precio_total__gt=0.00
                ).select_related("proveedor")

    context = {
        "ordenes_de_pago": ordenes_de_pago,
        "ordenventa": ordenventa,
        "fecha": fecha,  # Usa la variable fecha directamente
    }
    return render(request, "ver_orden_pago.html", context)


### Ordenes de cobranza :UUUUU
# usaremos una clase para eso pra probar igual no me pagan por innovar
# prbe y resutla mas facil hacerlo con una funcion talvez?? necesito mas exp para saber cual es mejor por ahora solo def
def ver_cobros_orden_venta(request, ordenventa_id):
    orden_venta = get_object_or_404(OrdenVenta, id=ordenventa_id)
    cobros = CobrosOrdenVenta.objects.filter(orden_venta=orden_venta)

    return render(request, 'cobros/lista_cobros_orden_venta.html', {
        'orden_venta': orden_venta,
        'cobros': cobros
    })


## vista automatica de facturas okokokok
def ver_facturas_por_ordenventa(request, ordenventa_id):
    # Buscar la orden de venta específica
    orden_venta = get_object_or_404(OrdenVenta, pk=ordenventa_id)

    # Buscar todas las facturas asociadas a esta orden de venta
    facturas = FacturaElectronica.objects.filter(orden_venta=orden_venta)

    # Pasar las facturas al contexto para ser mostradas en la plantilla
    context = {
        'orden_venta': orden_venta,
        'facturas': facturas,
    }

    return render(request, 'cobros/ver_facturas.html', context)



def eliminar_factura(request, factura_id):
    factura = get_object_or_404(FacturaElectronica, pk=factura_id)
    orden_venta_id = factura.orden_venta.id
    factura.delete()
    return redirect(reverse('ver_facturas_por_ordenventa', kwargs={'ordenventa_id': orden_venta_id}))


## editar factura:
def editar_factura(request, factura_id):
    factura = get_object_or_404(FacturaElectronica, pk=factura_id)
    if request.method == 'POST':
        form = FacturaElectronicaForm(request.POST, instance=factura)
        if form.is_valid():
            form.save()
            # Si usas HTMX, puedes querer devolver un fragmento actualizado en lugar de hacer una redirección completa
            return HttpResponse(render_to_string('fragmentos/factura_actualizada.html', {'factura': factura}))
    else:
        form = FacturaElectronicaForm(instance=factura)
    
    return HttpResponse(render_to_string('fragmentos/formulario_editar_factura.html', {'form': form, 'factura_id': factura.id}))

#Subir pdf ok para facturas  ok oko ok ok oko  
def upload_pdf_cobro(request, ordenventa_id):
    orden_venta = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    if request.method == "POST":
        pdf_file = request.FILES.get('pdf_file')
        if pdf_file:
            texto_completo = ""
            with pdfplumber.open(pdf_file) as pdf:
                for pagina in pdf.pages:
                    texto_completo += pagina.extract_text() + "\n"
            
            # Aquí incluyes todo el código de extracción que has definido anteriormente
            # Asegúrate de ajustar el manejo de los datos según sea necesario
            # Por ejemplo, asegúrate de convertir las cadenas de fechas y decimales correctamente
            # Buscar la serie y el correlativo
            patron_serie_correlativo = r"E\d{3}-\d+"
            resultado_serie_correlativo = re.search(patron_serie_correlativo, texto_completo)
            serie_correlativo = resultado_serie_correlativo.group() if resultado_serie_correlativo else "No encontrado"

            # Buscar la fecha de emisión
            patron_fecha_emision = r"Fecha de Emisión\s*:\s*(\d{2}/\d{2}/\d{4})"
            resultado_fecha_emision = re.search(patron_fecha_emision, texto_completo)
            fecha_emision = resultado_fecha_emision.group(1) if resultado_fecha_emision else "No encontrado"

            # Buscar el cliente (Señor(es))
            patron_cliente = r"Señor\(es\) :\s*(.*)\n"
            resultado_cliente = re.search(patron_cliente, texto_completo)
            cliente = resultado_cliente.group(1) if resultado_cliente else "No encontrado"

            # Buscar el RUC del cliente
            patron_ruc_cliente = r"RUC :\s*(\d+)"
            resultado_ruc_cliente = re.search(patron_ruc_cliente, texto_completo)
            ruc_cliente = resultado_ruc_cliente.group(1) if resultado_ruc_cliente else "No encontrado"

            # Buscar el tipo de moneda
            patron_tipo_moneda = r"Tipo de Moneda\s*:\s*(.*)\n"
            resultado_tipo_moneda = re.search(patron_tipo_moneda, texto_completo)
            tipo_moneda = resultado_tipo_moneda.group(1) if resultado_tipo_moneda else "No encontrado"

            # Ajustar la búsqueda de la descripción para que termine antes de los detalles financieros
            # Suponiendo que "Segun OS" marca el final relevante de la descripción del servicio
            patron_descripcion = r"Descripción\s*(.*?)(?=Sub Total Ventas|$)"
            resultado_descripcion = re.search(patron_descripcion, texto_completo, re.DOTALL)
            descripcion = resultado_descripcion.group(1).strip() if resultado_descripcion else "No encontrado"


            # Buscar el importe total
            patron_importe_total = r"Importe Total\s*:\s*\$(.*)\n"
            resultado_importe_total = re.search(patron_importe_total, texto_completo)
            importe_total = resultado_importe_total.group(1) if resultado_importe_total else "No encontrado"

            # Buscar el monto de la detracción
            patron_detraccion = r"Monto detracción:\s*S/\s*(.*)\n"
            resultado_detraccion = re.search(patron_detraccion, texto_completo)
            detraccion = resultado_detraccion.group(1) if resultado_detraccion else "0"

            # Buscar el monto neto a cobrar
            patron_neto_cobrar = r"Monto neto pendiente de pago\s*:\s*\$(.*)\n"
            resultado_neto_cobrar = re.search(patron_neto_cobrar, texto_completo)
            neto_cobrar = resultado_neto_cobrar.group(1) if resultado_neto_cobrar else "No encontrado"

            # Buscar el total de cuotas
            patron_total_cuotas = r"Total de Cuotas\s*:\s*(\d+)"
            resultado_total_cuotas = re.search(patron_total_cuotas, texto_completo)
            total_cuotas = resultado_total_cuotas.group(1) if resultado_total_cuotas else "No encontrado"

            # Buscar la fecha de vencimiento de la cuota
            patron_fecha_vencimiento = r"Fec. Venc.\s*Monto\n\d+\s*(\d{2}/\d{2}/\d{4})"
            resultado_fecha_vencimiento = re.search(patron_fecha_vencimiento, texto_completo)
            fecha_vencimiento = resultado_fecha_vencimiento.group(1) if resultado_fecha_vencimiento else "No encontrado"
         
         
         
            # Convertir fecha de emisión y fecha de vencimiento al formato correcto
            fecha_emision_format = datetime.strptime(fecha_emision, '%d/%m/%Y').strftime('%Y-%m-%d')
            fecha_vencimiento_format = datetime.strptime(fecha_vencimiento, '%d/%m/%Y').strftime('%Y-%m-%d')


            # Luego, creas una instancia de tu modelo con los datos extraídos
            factura = FacturaElectronica(
                orden_venta=orden_venta,
                serie_correlativo=serie_correlativo,
                fecha_emision=fecha_emision_format,
                cliente=cliente,
                ruc_cliente=ruc_cliente,
                tipo_moneda=tipo_moneda,
                descripcion=descripcion,
                importe_total=Decimal(importe_total.replace(',', '')),
                detraccion=Decimal(detraccion.replace(',', '')),
                monto_neto_cobrar=Decimal(neto_cobrar.replace(',', '')),
                total_cuotas=int(total_cuotas),
                fecha_vencimiento=fecha_vencimiento_format,
            )
            factura.save()
            
            # Redirigir o mostrar un mensaje de éxito
            return redirect(reverse('ver_facturas_por_ordenventa', args=[ordenventa_id])) # Asegúrate de definir esta URL
    
    return render(request, 'cobros/upload_pdf_cobros.html')






















# desde aqui no hay nada interesante :P


# Vistas para OrdenVenta
class ListaOrdenesVenta(generics.ListCreateAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


class DetalleOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


# Vistas para ItemOrdenVenta
class ListaItemsOrdenVenta(generics.ListCreateAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer


class DetalleItemOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer
